import re
from copy import deepcopy
from openpyxl import load_workbook, Workbook


INPUT_PATH = r"public/excel/Air_India_Pilots.xlsx"
OUTPUT_PATH = r"public/excel/Air_India_Pilots_updated.xlsx"


PHONE_SPLIT_RE = re.compile(r"\s*/\s*")


def normalize_phone(s: str) -> str:
    digits = re.sub(r"\D+", "", s or "")
    return digits


def is_blank_row(values) -> bool:
    for v in values:
        if v is None:
            continue
        if isinstance(v, str) and v.strip() == "":
            continue
        return False
    return True


def find_phone_col(headers):
    candidates = []
    for idx, h in enumerate(headers):
        hs = (str(h).strip().lower() if h is not None else "")
        if not hs:
            continue
        score = 0
        if "phone" in hs or "mobile" in hs:
            score += 3
        if "number" in hs or "contact" in hs or "whatsapp" in hs:
            score += 2
        if score:
            candidates.append((score, idx))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    return candidates[0][1]


def main():
    wb = load_workbook(INPUT_PATH)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    phone_col = find_phone_col(headers)
    if phone_col is None:
        raise SystemExit(
            f"Could not detect phone column from headers: {headers}"
        )

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = ws.title
    out_ws.append(headers)

    written = 0
    duplicated = 0
    removed_blank = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if is_blank_row(row):
            removed_blank += 1
            continue

        row_list = list(row)
        phone_val = row_list[phone_col]
        if phone_val is None or (isinstance(phone_val, str) and phone_val.strip() == ""):
            out_ws.append(row_list)
            written += 1
            continue

        phone_str = str(phone_val).strip()
        parts = [p for p in PHONE_SPLIT_RE.split(phone_str) if p and p.strip()]
        if len(parts) <= 1:
            out_ws.append(row_list)
            written += 1
            continue

        # Create one row per phone
        for p in parts:
            new_row = deepcopy(row_list)
            new_row[phone_col] = normalize_phone(p)
            out_ws.append(new_row)
            written += 1
            duplicated += 1

    out_wb.save(OUTPUT_PATH)
    print("Saved:", OUTPUT_PATH)
    print("phone_col_index:", phone_col, "header:", headers[phone_col])
    print("written_rows:", written)
    print("split_rows_created:", duplicated)
    print("blank_rows_removed:", removed_blank)


if __name__ == "__main__":
    main()

